from rest_framework import viewsets, filters
from rest_framework.permissions import AllowAny
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from schools.models import School
from .models import ClassRoom, Section, Subject, StudentProfile, TeacherAssignment
from .serializers import (
    SchoolSerializer, ClassRoomSerializer, SectionSerializer,
    SubjectSerializer, StudentProfileSerializer, TeacherAssignmentSerializer
)


class SchoolListAPI(APIView):
    def get(self, request):
        schools = School.objects.all()
        serializer = SchoolSerializer(schools, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ClassRoomViewSet(viewsets.ModelViewSet):
    queryset = ClassRoom.objects.select_related('school').prefetch_related('sections', 'students').all()
    serializer_class = ClassRoomSerializer
    permission_classes = [AllowAny]  # TEMP: dev-only open access
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['school']
    search_fields = ['name']
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get class summary with student counts"""
        school_id = request.query_params.get('school')
        if not school_id:
            return Response({"detail": "school parameter required"}, status=status.HTTP_400_BAD_REQUEST)
        
        classrooms = ClassRoom.objects.filter(school_id=school_id).prefetch_related('students')
        data = []
        for classroom in classrooms:
            data.append({
                'id': classroom.id,
                'name': classroom.name,
                'description': classroom.description,
                'student_count': classroom.students.count(),
                'subject_count': Subject.objects.filter(school_id=school_id).count()  # subjects are school-wide
            })
        return Response(data)
    
    @action(detail=True, methods=['get'])
    def students(self, request, pk=None):
        """Get all students in a specific class"""
        classroom = self.get_object()
        students = StudentProfile.objects.filter(classroom=classroom).select_related('user', 'guardian', 'section')
        serializer = StudentProfileSerializer(students, many=True, context={'request': request})
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def subjects(self, request, pk=None):
        """Get all subjects for a specific class with assigned teachers"""
        classroom = self.get_object()
        # Get all subjects for this school
        subjects = Subject.objects.filter(school=classroom.school)
        
        result = []
        for subject in subjects:
            # Get teachers assigned to this subject and class
            assignments = TeacherAssignment.objects.filter(
                subject=subject,
                classroom=classroom
            ).select_related('teacher')
            
            teachers = []
            for assignment in assignments:
                teacher = assignment.teacher
                teachers.append({
                    'id': teacher.id,
                    'name': f"{teacher.first_name} {teacher.last_name}".strip() or teacher.username,
                    'username': teacher.username
                })
            
            result.append({
                'id': subject.id,
                'name': subject.name,
                'code': subject.code,
                'teachers': teachers,
                'notifications': 0  # Placeholder for future notification count
            })
        
        return Response(result)


class SectionViewSet(viewsets.ModelViewSet):
    queryset = Section.objects.select_related('classroom__school').all()
    serializer_class = SectionSerializer
    permission_classes = [AllowAny]  # TEMP: dev-only open access
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['classroom']


class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.select_related('school').all()
    serializer_class = SubjectSerializer
    permission_classes = [AllowAny]  # TEMP: dev-only open access
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['school']
    search_fields = ['name','code']
    
    @action(detail=True, methods=['get'])
    def detail(self, request, pk=None):
        """Get detailed subject information including assignments, results, and attendance"""
        subject = self.get_object()
        
        # Get teacher assignments
        assignments = TeacherAssignment.objects.filter(subject=subject).select_related(
            'teacher', 'classroom', 'section'
        )
        
        assignments_data = []
        for assignment in assignments:
            teacher = assignment.teacher
            assignments_data.append({
                'id': assignment.id,
                'teacher': {
                    'id': teacher.id,
                    'name': f"{teacher.first_name} {teacher.last_name}".strip() or teacher.username,
                    'username': teacher.username
                },
                'classroom': {
                    'id': assignment.classroom.id,
                    'name': assignment.classroom.name
                },
                'section': {
                    'id': assignment.section.id,
                    'name': assignment.section.name
                } if assignment.section else None
            })
        
        # Get recent results for this subject
        from results.models import Result
        recent_results = Result.objects.filter(subject=subject).select_related(
            'examination', 'student__user'
        ).order_by('-examination__exam_date')[:20]
        
        results_data = []
        for result in recent_results:
            results_data.append({
                'examination': result.examination.name,
                'student': f"{result.student.user.first_name} {result.student.user.last_name}".strip() or result.student.user.username,
                'total_obtained': float(result.total_obtained),
                'grade': result.grade
            })
        
        return Response({
            'id': subject.id,
            'name': subject.name,
            'code': subject.code,
            'assignments': assignments_data,
            'recent_results': results_data,
            'total_assignments': len(assignments_data),
            'notifications': 0  # Placeholder
        })


class StudentProfileViewSet(viewsets.ModelViewSet):
    queryset = StudentProfile.objects.select_related('user', 'school', 'classroom', 'section', 'guardian').all()
    serializer_class = StudentProfileSerializer
    permission_classes = [AllowAny]  # TEMP: dev-only open access
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['classroom__school', 'school', 'classroom', 'section', 'guardian']
    search_fields = ['user__username','user__first_name','user__last_name','roll_number']
    parser_classes = [MultiPartParser, FormParser]
    
    @action(detail=True, methods=['get'])
    def detail(self, request, pk=None):
        """Get detailed student information including results and attendance"""
        student = self.get_object()
        
        # Basic student info
        from users.serializers import UserSerializer
        user_data = UserSerializer(student.user, context={'request': request}).data
        
        # Get recent results
        from results.models import Result
        recent_results = Result.objects.filter(student=student).select_related(
            'examination', 'subject'
        ).order_by('-examination__exam_date')[:10]
        
        results_data = []
        for result in recent_results:
            results_data.append({
                'examination': result.examination.name,
                'subject': result.subject.name,
                'total_obtained': float(result.total_obtained),
                'grade': result.grade,
                'gpa': float(result.gpa),
                'is_passed': result.is_passed
            })
        
        # Get attendance summary (placeholder - implement when attendance app is ready)
        attendance_summary = {
            'total_days': 0,
            'present_days': 0,
            'absent_days': 0,
            'percentage': 0
        }
        
        return Response({
            'id': student.id,
            'user': user_data,
            'classroom': {'id': student.classroom.id, 'name': student.classroom.name} if student.classroom else None,
            'section': {'id': student.section.id, 'name': student.section.name} if student.section else None,
            'roll_number': student.roll_number,
            'guardian_name': student.guardian_name,
            'guardian': UserSerializer(student.guardian, context={'request': request}).data if student.guardian else None,
            'recent_results': results_data,
            'attendance': attendance_summary
        })
    
    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_photo(self, request, pk=None):
        """Upload photo for a student"""
        student = self.get_object()
        
        if 'photo' in request.FILES:
            student.user.photo = request.FILES['photo']
            student.user.save()
            
            from users.serializers import UserSerializer
            return Response({
                "message": "Photo uploaded successfully",
                "user": UserSerializer(student.user, context={'request': request}).data
            })
        
        return Response({"error": "No photo provided"}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export students to CSV"""
        import csv
        from django.http import HttpResponse
        
        school_id = request.query_params.get('school')
        classroom_id = request.query_params.get('classroom')
        
        qs = self.filter_queryset(self.get_queryset())
        
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="students_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Serial', 'Student Name', 'Username', 'Class', 'Section', 'Roll Number', 'Parent Name', 'Parent Username'])
        
        for idx, sp in enumerate(qs, start=1):
            user = sp.user
            student_name = f"{user.first_name} {user.last_name}".strip() or user.username
            classroom = sp.classroom.name if sp.classroom else ''
            section = sp.section.name if sp.section else ''
            parent_name = sp.guardian_name or ''
            parent_username = sp.guardian.username if sp.guardian else ''
            
            writer.writerow([idx, student_name, user.username, classroom, section, sp.roll_number or '', parent_name, parent_username])
        
        return response


class TeacherAssignmentViewSet(viewsets.ModelViewSet):
    queryset = TeacherAssignment.objects.select_related('teacher','subject','classroom','section').all()
    serializer_class = TeacherAssignmentSerializer
    permission_classes = [AllowAny]  # TEMP: dev-only open access
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['classroom__school', 'teacher']
    search_fields = ['teacher__username','teacher__first_name','teacher__last_name']


# ---- Import Students API ----
from django.contrib.auth import get_user_model
from django.db import transaction
import io
import csv

User = get_user_model()


class ImportStudentsAPI(APIView):
    permission_classes = [AllowAny]  # You may tighten later

    REQUIRED_COLUMNS = {"username", "first_name", "last_name", "classroom", "section", "roll_number"}

    def post(self, request):
        school_id = request.POST.get("school") or request.query_params.get("school")
        if not school_id:
            return Response({"detail": "Parameter 'school' is required."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            school = School.objects.get(pk=school_id)
        except School.DoesNotExist:
            return Response({"detail": "Invalid school id."}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file uploaded. Use form field 'file'."}, status=status.HTTP_400_BAD_REQUEST)

        name = file.name.lower()
        try:
            if name.endswith(".csv"):
                created, updated, errors = self._import_csv(file, school)
            elif name.endswith(".docx"):
                created, updated, errors = self._import_docx(file, school)
            elif name.endswith(".pdf"):
                created, updated, errors = self._import_pdf(file, school)
            elif name.endswith((".xlsx", ".xlsm")):
                created, updated, errors = self._import_xlsx(file, school)
            elif name.endswith((".png", ".jpg", ".jpeg")):
                created, updated, errors = self._import_image(file, school)
            else:
                return Response({"detail": "Unsupported file type. Use CSV, DOCX, PDF, or image."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"detail": f"Failed to import: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            "message": "Import complete",
            "created": created,
            "updated": updated,
            "errors": errors,
        }, status=status.HTTP_200_OK)

    def _normalize_headers(self, headers):
        return [h.strip().lower().replace(" ", "_") for h in headers]

    @transaction.atomic
    def _import_csv(self, uploaded_file, school):
        created, updated = 0, 0
        errors = []
        # Attempt utf-8-sig then fallback latin-1
        content = uploaded_file.read()
        for enc in ["utf-8-sig", "utf-8", "latin-1"]:
            try:
                text = content.decode(enc)
                break
            except Exception:
                continue
        else:
            raise ValueError("Unable to decode CSV file")

        reader = csv.DictReader(io.StringIO(text))
        reader.fieldnames = self._normalize_headers(reader.fieldnames or [])

        # Optional: allow partial columns
        supported = {"username", "first_name", "last_name", "password", "classroom", "section", "roll_number",
                     "parent", "guardian", "guardian_name", "father_name", "mother_name"}

        row_num = 1
        for row in reader:
            row_num += 1
            data = {k: (row.get(k) or "").strip() for k in supported}
            try:
                cu, uu = self._create_or_update_student(data, school)
                created += cu
                updated += uu
            except Exception as e:
                errors.append({"row": row_num, "error": str(e)})

        return created, updated, errors

    def _import_docx(self, uploaded_file, school):
        try:
            import docx  # python-docx
        except Exception:
            raise ValueError("python-docx not installed on server")
        # Very simple extraction: read table rows into dicts with header
        document = docx.Document(uploaded_file)
        created, updated, errors = 0, 0, []
        for table in document.tables:
            headers = self._normalize_headers([cell.text for cell in table.rows[0].cells]) if table.rows else []
            supported = {"username", "first_name", "last_name", "password", "classroom", "section", "roll_number",
                         "parent", "guardian", "guardian_name", "father_name", "mother_name"}
            for r_i, row in enumerate(table.rows[1:], start=2):
                values = [cell.text.strip() for cell in row.cells]
                data = {h: (values[i] if i < len(values) else "") for i, h in enumerate(headers) if h in supported}
                try:
                    cu, uu = self._create_or_update_student(data, school)
                    created += cu
                    updated += uu
                except Exception as e:
                    errors.append({"row": r_i, "error": str(e)})
        return created, updated, errors

    def _import_pdf(self, uploaded_file, school):
        # Try pdfplumber for table extraction
        try:
            import pdfplumber
        except Exception:
            raise ValueError("pdfplumber not installed on server")
        created, updated, errors = 0, 0, []
        with pdfplumber.open(uploaded_file) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables() or []
                for t in tables:
                    if not t or not t[0]:
                        continue
                    headers = self._normalize_headers(t[0])
                    supported = {"username", "first_name", "last_name", "password", "classroom", "section", "roll_number",
                                 "parent", "guardian", "guardian_name", "father_name", "mother_name"}
                    for r_i, row in enumerate(t[1:], start=2):
                        data = {h: (row[i].strip() if i < len(row) and row[i] else "") for i, h in enumerate(headers) if h in supported}
                        try:
                            cu, uu = self._create_or_update_student(data, school)
                            created += cu
                            updated += uu
                        except Exception as e:
                            errors.append({"row": r_i, "error": str(e)})
        return created, updated, errors

    def _import_xlsx(self, uploaded_file, school):
        try:
            from openpyxl import load_workbook
        except Exception:
            raise ValueError("openpyxl not installed on server")

        created, updated, errors = 0, 0, []
        try:
            wb = load_workbook(uploaded_file, data_only=True)
        except Exception as e:
            raise ValueError(f"Failed to open Excel file: {e}")

        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            return 0, 0, [{"row": 0, "error": "Empty Excel sheet"}]
        headers = self._normalize_headers([str(h) if h is not None else '' for h in headers])
        supported = {"username", "first_name", "last_name", "password", "classroom", "section", "roll_number",
                    "parent", "guardian", "guardian_name", "father_name", "mother_name"}

        for idx, row in enumerate(rows_iter, start=2):
            values = [str(v).strip() if v is not None else '' for v in row]
            data = {h: (values[i] if i < len(values) else '') for i, h in enumerate(headers) if h in supported}
            try:
                cu, uu = self._create_or_update_student(data, school)
                created += cu
                updated += uu
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})
        return created, updated, errors

    def _import_image(self, uploaded_file, school):
        """
        Detect table grid and OCR per-cell to map 6 columns:
        [serial, parent, student, class, section, roll_number]
        Uses OpenCV (cv2) + Tesseract (ben+eng). Falls back to naive OCR if grid detection fails.
        """
        try:
            from PIL import Image
            import pytesseract
            # Set Tesseract path for Windows
            import os
            if os.name == 'nt':  # Windows
                tesseract_path = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
                if os.path.exists(tesseract_path):
                    pytesseract.pytesseract.tesseract_cmd = tesseract_path
        except Exception:
            raise ValueError("pillow/pytesseract not installed on server")

        # Try advanced grid-based extraction
        try:
            import cv2
            import numpy as np
        except Exception:
            cv2 = None
            np = None

        created, updated, errors = 0, 0, []

        # Load image
        image = Image.open(uploaded_file).convert('RGB')

        def ocr_cell(img_pil):
            try:
                # Bengali + English; page segmentation mode 7 (single text line) works well for cells
                return pytesseract.image_to_string(
                    img_pil,
                    lang='ben+eng',
                    config='--psm 7'
                ).strip()
            except Exception:
                return ''

        used_advanced = False
        if cv2 is not None and np is not None:
            try:
                # Convert to OpenCV format
                img_cv = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
                gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
                # Adaptive threshold for robust binarization
                thr = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C,
                                            cv2.THRESH_BINARY_INV, 15, 10)

                # Detect horizontal and vertical lines
                h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
                v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))
                h_lines = cv2.morphologyEx(thr, cv2.MORPH_OPEN, h_kernel, iterations=2)
                v_lines = cv2.morphologyEx(thr, cv2.MORPH_OPEN, v_kernel, iterations=2)
                table_mask = cv2.add(h_lines, v_lines)

                # Find contours of boxes (cells)
                contours, _ = cv2.findContours(table_mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
                boxes = []
                for cnt in contours:
                    x, y, w, h = cv2.boundingRect(cnt)
                    # Filter too small or too large boxes
                    if w < 40 or h < 20:
                        continue
                    boxes.append((y, x, w, h))
                if not boxes:
                    raise RuntimeError('No boxes detected')

                # Sort boxes top-to-bottom, then left-to-right
                boxes.sort()
                # Group boxes into rows by y proximity
                rows = []
                row = []
                last_y = None
                tol = 12  # vertical tolerance
                for (y, x, w, h) in boxes:
                    if last_y is None or abs(y - last_y) <= tol:
                        row.append((y, x, w, h))
                        last_y = y
                    else:
                        rows.append(sorted(row, key=lambda b: b[1]))
                        row = [(y, x, w, h)]
                        last_y = y
                if row:
                    rows.append(sorted(row, key=lambda b: b[1]))

                # Heuristic: skip the top header row(s) by requiring first cell to be a number
                processed_rows = 0
                for r_idx, r in enumerate(rows, start=1):
                    # Expect at least 6 columns; if more, we take first 6
                    if len(r) < 6:
                        continue
                    cells = r[:6]
                    texts = []
                    for (y, x, w, h) in cells:
                        crop = img_cv[y:y+h, x:x+w]
                        crop = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)
                        crop_pil = Image.fromarray(crop)
                        texts.append(ocr_cell(crop_pil))

                    # Validate first cell numeric
                    serial = ''.join(ch for ch in texts[0] if ch.isdigit())
                    if not serial:
                        # likely header
                        continue

                    parent = texts[1]
                    student = texts[2]
                    classroom = texts[3]
                    section = texts[4]
                    roll_number = ''.join(ch for ch in texts[5] if ch.isdigit()) or texts[5]

                    # Map to creation data
                    data = {
                        # We'll set username empty to auto-generate from student
                        "username": "",
                        "first_name": student.split()[0] if student else "",
                        "last_name": ' '.join(student.split()[1:]) if student and len(student.split()) > 1 else "",
                        "password": "",
                        "classroom": classroom,
                        "section": section,
                        "roll_number": roll_number,
                        "guardian_name": parent,
                    }
                    try:
                        cu, uu = self._create_or_update_student(data, school)
                        created += cu
                        updated += uu
                        processed_rows += 1
                    except Exception as e:
                        errors.append({"row": r_idx, "error": str(e)})

                if processed_rows == 0:
                    raise RuntimeError('No data rows recognized from table')

                used_advanced = True
            except Exception as e:
                # Fall back to naive OCR parsing
                used_advanced = False

        if not used_advanced:
            # Fallback: naive whole-image OCR and parse by multiple spaces or commas
            try:
                text = pytesseract.image_to_string(image, lang='ben+eng', config='--psm 6')
            except Exception as e:
                raise ValueError(f"OCR failed: {e}")

            lines = [l for l in (text.splitlines()) if l and len(l.strip()) > 0]
            # Try to find the first line with a leading number to start data rows
            started = False
            row_idx = 0
            for l in lines:
                parts = [p for p in l.strip().split('\t') if p]  # sometimes tesseract uses tabs
                if len(parts) < 2:
                    # split by 2+ spaces
                    parts = [p for p in filter(None, [p.strip() for p in __import__('re').split(r"\s{2,}", l)])]
                # Expect at least 6 columns; if more, keep first 6
                if len(parts) >= 2 and (parts[0].strip().isdigit() or parts[0].strip().replace('.', '').isdigit()):
                    started = True
                if not started:
                    continue
                row_idx += 1
                if len(parts) < 6:
                    # can't parse; keep as error
                    errors.append({"row": row_idx, "error": f"Unparsable row: {l}"})
                    continue
                cols = parts[:6]
                serial = cols[0]
                parent = cols[1]
                student = cols[2]
                classroom = cols[3]
                section = cols[4]
                roll_number = ''.join(ch for ch in cols[5] if ch.isdigit()) or cols[5]

                data = {
                    "username": "",
                    "first_name": student.split()[0] if student else "",
                    "last_name": ' '.join(student.split()[1:]) if student and len(student.split()) > 1 else "",
                    "password": "",
                    "classroom": classroom,
                    "section": section,
                    "roll_number": roll_number,
                    "guardian_name": parent,
                }
                try:
                    cu, uu = self._create_or_update_student(data, school)
                    created += cu
                    updated += uu
                except Exception as e:
                    errors.append({"row": row_idx, "error": str(e)})

        return created, updated, errors

    def _create_or_update_student(self, data, school):
        """
        Create or update a student and StudentProfile. Returns (created_count, updated_count)
        Expected fields in data: username, first_name, last_name, password, classroom, section, roll_number,
        and optional guardian fields (guardian_name/parent/father_name/mother_name).
        Also auto-creates a Parent user+Profile when guardian_name is provided.
        """
        username = data.get("username") or ""
        first_name = data.get("first_name") or ""
        last_name = data.get("last_name") or ""
        password = data.get("password") or ""
        classroom_name = data.get("classroom") or ""
        section_name = data.get("section") or ""
        roll_number = data.get("roll_number") or ""
        guardian_name = (
            data.get("guardian_name") or data.get("guardian") or data.get("parent")
            or data.get("father_name") or data.get("mother_name") or ""
        ).strip()

        if not username and not first_name:
            raise ValueError("Either username or first_name is required")

        # Ensure username
        if not username:
            base = (first_name or "student").lower().replace(" ", "")
            candidate = base
            idx = 1
            while User.objects.filter(username=candidate).exists():
                idx += 1
                candidate = f"{base}{idx}"
            username = candidate

        user, created_user = User.objects.get_or_create(username=username, defaults={
            "first_name": first_name,
            "last_name": last_name,
        })
        # If user existed, optionally update names
        if not created_user:
            changed = False
            if first_name and user.first_name != first_name:
                user.first_name = first_name; changed = True
            if last_name and user.last_name != last_name:
                user.last_name = last_name; changed = True
            if changed:
                user.save()
        if password:
            user.set_password(password)
            user.save()

        # Classroom and Section
        classroom = None
        if classroom_name:
            classroom, _ = ClassRoom.objects.get_or_create(school=school, name=classroom_name)
        section = None
        if classroom and section_name:
            section, _ = Section.objects.get_or_create(classroom=classroom, name=section_name)

        # Prepare guardian user if provided
        guardian_user = None
        if guardian_name:
            # Create or reuse a guardian user by generated username
            base = guardian_name.replace(' ', '').lower() or 'parent'
            candidate = base
            idx = 1
            from django.contrib.auth import get_user_model
            U = get_user_model()
            while U.objects.filter(username=candidate).exists():
                idx += 1
                candidate = f"{base}{idx}"
            guardian_user, created_g = U.objects.get_or_create(username=candidate, defaults={
                "first_name": guardian_name.split()[0] if guardian_name else "",
                "last_name": ' '.join(guardian_name.split()[1:]) if guardian_name and len(guardian_name.split()) > 1 else "",
            })
            # Ensure Profile with role parent
            from users.models import Profile as UserProfile
            UserProfile.objects.update_or_create(user=guardian_user, defaults={"school": school, "role": "parent"})

        sp, created_profile = StudentProfile.objects.get_or_create(user=user, defaults={
            "school": school,
            "classroom": classroom,
            "section": section,
            "roll_number": roll_number or None,
            "guardian_name": guardian_name or None,
            "guardian": guardian_user,
        })
        if not created_profile:
            changed = False
            if sp.school_id != school.id:
                sp.school = school; changed = True
            if classroom and sp.classroom_id != classroom.id:
                sp.classroom = classroom; changed = True
            if section and (sp.section_id or None) != (section.id if section else None):
                sp.section = section; changed = True
            if roll_number and sp.roll_number != roll_number:
                sp.roll_number = roll_number; changed = True
            if guardian_name and sp.guardian_name != guardian_name:
                sp.guardian_name = guardian_name; changed = True
            if guardian_user and (sp.guardian_id or None) != guardian_user.id:
                sp.guardian = guardian_user; changed = True
            if changed:
                sp.save()
            return 0, 1

        return 1, 0
